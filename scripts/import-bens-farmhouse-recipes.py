#!/usr/bin/env python3
"""
Import Ben's Farmhouse recipes from Excel into Supabase recipes table.
Assigns all recipes to owner_email: meerim@bensfarmhouse.com
"""

import openpyxl
import json
import urllib.request
import urllib.error
import time
import random
import string
import sys

# Supabase config
SUPABASE_URL = "https://sijhdlxagliknvhkosjo.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNpamhkbHhhZ2xpa252aGtvc2pvIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MDQ0NDYzMCwiZXhwIjoyMDc2MDIwNjMwfQ.RurV1sZnGm80--rTl5uXG3UoMc061G8fb2ydjmJXCA0"
OWNER_EMAIL = "meerim@bensfarmhouse.com"

EXCEL_PATH = "/Users/talalbazerbachi/Downloads/ben's-farmhouse-2026-02-10-08_22-with-ingredients.xlsx"


def generate_recipe_id():
    """Generate a recipe ID in the same format as the app: recipe_{timestamp}_{random}"""
    ts = int(time.time() * 1000)
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=9))
    return f"recipe_{ts}_{rand}"


def parse_excel(filepath):
    """Parse the Excel file and extract all recipes with their ingredients.

    Strategy: Two-pass approach.
    Pass 1: Find all "Recipe ID" rows to identify where recipes start.
    Pass 2: For each recipe, look back to find the recipe name and category.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Recipes']
    max_row = ws.max_row

    # Pass 1: Find all rows that contain "Recipe ID" in column B
    recipe_id_rows = []
    for row in range(1, max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        if b_val and 'Recipe ID' in str(b_val):
            recipe_id_rows.append(row)

    print(f"Found {len(recipe_id_rows)} recipe ID markers")

    # Pass 2: For each recipe ID row, find the recipe name (text-only in col A
    # immediately before) and category (text-only in col A before that)
    recipes = []
    current_category = None

    for idx, rid_row in enumerate(recipe_id_rows):
        # Look backwards from rid_row to find the recipe name
        # The recipe name is the closest row above with text only in col A
        recipe_name = None
        recipe_name_row = None

        for r in range(rid_row - 1, max(rid_row - 5, 0), -1):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value
            if a_val and not b_val:
                val = str(a_val).strip()
                # Skip rows that are clearly not names
                if val in ["BEN'S FARMHOUSE", "Ingredients", "S.No"] or val == '':
                    continue
                if '20' in val and ('-' in val or '/' in val) and len(val) < 25:
                    continue  # date
                recipe_name = val
                recipe_name_row = r
                break

        if not recipe_name:
            continue

        # Now check if there's a category above the recipe name
        # Look backwards from recipe_name_row for another text-only row
        # that isn't part of the previous recipe's ingredients
        for r in range(recipe_name_row - 1, max(recipe_name_row - 5, 0), -1):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value

            if a_val is None and b_val is None:
                continue  # skip empty rows

            if a_val and not b_val:
                val = str(a_val).strip()
                # Skip headers and dates
                if val in ["BEN'S FARMHOUSE", "Ingredients", "S.No"] or val == '':
                    break
                if '20' in val and ('-' in val or '/' in val) and len(val) < 25:
                    break

                # Check if this is a category (not the previous recipe's last ingredient)
                # A category row has text only in col A and is NOT a number (ingredient S.No)
                try:
                    float(val)
                    break  # It's a number (ingredient row) - stop looking
                except ValueError:
                    pass

                # Check: is this row between the end of the previous recipe and this recipe?
                # If the previous recipe's ingredients ended before this row, it's a category
                if idx > 0:
                    prev_rid_row = recipe_id_rows[idx - 1]
                    # Previous recipe's ingredients end somewhere after prev_rid_row
                    # If this row is after the previous recipe's ingredient area, it's a category
                    if r > prev_rid_row:
                        current_category = val
                else:
                    # First recipe - anything above is a category
                    current_category = val
                break
            else:
                # Row has data in col B - it's part of previous recipe
                break

        # Parse the recipe data
        recipe = parse_recipe(ws, recipe_name, rid_row, max_row, current_category)
        if recipe:
            recipes.append(recipe)

    return recipes


def parse_recipe(ws, recipe_name, rid_row, max_row, category):
    """Parse a single recipe given the recipe name and the row containing Recipe ID."""
    # Extract metadata from rid_row and following rows
    recipe_id_source = None
    total_yield = None
    total_calories = None
    per_serving_weight = None
    description = None
    prep_time = None

    # Parse metadata rows (Recipe ID row and a few rows below)
    for row in range(rid_row, min(rid_row + 8, max_row + 1)):
        b_val = ws.cell(row=row, column=2).value
        c_val = ws.cell(row=row, column=3).value

        if b_val:
            b_str = str(b_val).strip()
            if 'Recipe ID' in b_str:
                recipe_id_source = str(c_val).strip() if c_val else None
            elif 'Total Yield' in b_str:
                try:
                    total_yield = float(c_val) if c_val else None
                except (ValueError, TypeError):
                    total_yield = None
            elif 'Description' in b_str:
                description = str(c_val).strip() if c_val else ""
            elif 'Preparation Time' in b_str:
                try:
                    prep_time = int(float(c_val)) if c_val else None
                except (ValueError, TypeError):
                    prep_time = None

        # Check columns F,G for calories and per serving weight
        for col in [6, 7]:
            h = ws.cell(row=row, column=col).value
            if h:
                h_str = str(h).strip()
                if 'Calories' in h_str:
                    v = ws.cell(row=row, column=col + 1).value
                    if v is not None:
                        try:
                            total_calories = float(v)
                        except (ValueError, TypeError):
                            pass
                elif 'Per Serving Weight' in h_str:
                    v = ws.cell(row=row, column=col + 1).value
                    if v is not None:
                        try:
                            per_serving_weight = float(v)
                        except (ValueError, TypeError):
                            pass

        # Check if we've hit the Ingredients section
        a_val = ws.cell(row=row, column=1).value
        if a_val and str(a_val).strip() == 'Ingredients':
            break

    # Find Ingredients section
    ingredients_start = None
    for row in range(rid_row, min(rid_row + 12, max_row + 1)):
        a_val = ws.cell(row=row, column=1).value
        if a_val and str(a_val).strip() == 'Ingredients':
            # Next row should be S.No header
            next_row = row + 1
            next_a = ws.cell(row=next_row, column=1).value
            if next_a and 'S.No' in str(next_a):
                ingredients_start = next_row + 1
            else:
                ingredients_start = next_row
            break

    if ingredients_start is None:
        return None

    # Parse ingredients
    ingredients = []
    row = ingredients_start
    while row <= max_row:
        a_val = ws.cell(row=row, column=1).value

        if a_val is not None:
            try:
                int(float(str(a_val)))
            except (ValueError, TypeError):
                break  # Not a numbered row - end of ingredients

            ingredient_id = ws.cell(row=row, column=2).value
            ingredient_name = ws.cell(row=row, column=3).value
            ingredient_type = ws.cell(row=row, column=4).value
            sub_recipe_name = ws.cell(row=row, column=5).value
            quantity = ws.cell(row=row, column=6).value
            yield_pct = ws.cell(row=row, column=7).value
            yield_qty = ws.cell(row=row, column=8).value
            cost = ws.cell(row=row, column=9).value

            qty_val = 0
            try:
                qty_val = float(yield_qty) if yield_qty else (float(quantity) if quantity else 0)
            except (ValueError, TypeError):
                qty_val = 0

            ing = {
                "name": str(ingredient_name).strip() if ingredient_name else "",
                "unit": "G",
                "source": str(ingredient_type).strip() if ingredient_type else "Ingredient",
                "quantity": qty_val,
                "cost": float(cost) if cost else 0,
                "allergens": [],
                "yield_percent": float(yield_pct) if yield_pct else 100,
                "quantity_in_grams": qty_val,
                "yield_quantity_in_grams": qty_val,
                "ingredient_id": str(ingredient_id).strip() if ingredient_id else "",
                "entity_ingredient_id": f"imported_{generate_recipe_id()}",
                "is_starred": False,
                "user_allergens": [],
                "composite_ingredients": [],
                "may_contain_allergens": [],
                "user_may_contain_allergens": [],
                "nutrients": {}
            }

            if sub_recipe_name:
                ing["sub_recipe_name"] = str(sub_recipe_name).strip()

            ingredients.append(ing)
        else:
            # Empty col A - check if truly end of ingredients
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            if not b_val and not c_val:
                break

        row += 1

    # Calculate serving info
    serving_size = per_serving_weight if per_serving_weight else 100
    total_yield_val = total_yield if total_yield else sum(i.get('quantity_in_grams', 0) for i in ingredients)
    if total_yield_val == 0:
        total_yield_val = 100
    scale_factor = total_yield_val / serving_size if serving_size > 0 else 1

    recipe = {
        "id": generate_recipe_id(),
        "name": recipe_name,
        "category": category or "",
        "status": "IMPORTED",
        "description": description or "",
        "preparation_time_minutes": prep_time,
        "owner_email": OWNER_EMAIL,
        "serving": {
            "scale_factor": round(scale_factor, 2),
            "serving_unit": "g",
            "total_yield_grams": total_yield_val,
            "serving_size_grams": serving_size,
            "serving_description": "",
            "servings_per_package": 1,
            "total_cooked_weight_grams": total_yield_val
        },
        "diet_types": [],
        "allergens": [],
        "may_contain_allergens": [],
        "nutrition": {
            "summary": {
                "calories": total_calories or 0,
                "net_carbs": 0
            },
            "per_serving": {},
            "per_recipe_total": {},
            "total_yield_grams": total_yield_val,
            "serving_size_grams": serving_size
        },
        "ingredients": ingredients,
        "inventory": {
            "stock_unit": "portions",
            "stock_quantity": 0,
            "min_stock_alert": None,
            "last_stock_update": None
        },
        "_source_recipe_id": recipe_id_source
    }

    return recipe


def insert_recipes_batch(recipes, batch_size=50):
    """Insert recipes into Supabase in batches."""
    total = len(recipes)
    inserted = 0
    errors = 0

    for i in range(0, total, batch_size):
        batch = recipes[i:i + batch_size]

        # Clean up internal fields
        clean_batch = []
        for r in batch:
            clean = {k: v for k, v in r.items() if not k.startswith('_')}
            clean_batch.append(clean)

        data = json.dumps(clean_batch).encode('utf-8')

        url = f"{SUPABASE_URL}/rest/v1/recipes"
        req = urllib.request.Request(url, data=data, method='POST')
        req.add_header('apikey', SERVICE_ROLE_KEY)
        req.add_header('Authorization', f'Bearer {SERVICE_ROLE_KEY}')
        req.add_header('Content-Type', 'application/json')
        req.add_header('Prefer', 'return=minimal')

        try:
            resp = urllib.request.urlopen(req)
            status = resp.getcode()
            inserted += len(batch)
            print(f"  Batch {i//batch_size + 1}: Inserted {len(batch)} recipes (total: {inserted}/{total}) - Status: {status}")
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8')
            print(f"  Batch {i//batch_size + 1}: ERROR {e.code} - {error_body}")
            errors += len(batch)

            # Try one by one for failed batch
            print(f"  Retrying batch one by one...")
            for recipe in clean_batch:
                single_data = json.dumps(recipe).encode('utf-8')
                single_req = urllib.request.Request(url, data=single_data, method='POST')
                single_req.add_header('apikey', SERVICE_ROLE_KEY)
                single_req.add_header('Authorization', f'Bearer {SERVICE_ROLE_KEY}')
                single_req.add_header('Content-Type', 'application/json')
                single_req.add_header('Prefer', 'return=minimal')

                try:
                    urllib.request.urlopen(single_req)
                    inserted += 1
                    errors -= 1
                except urllib.error.HTTPError as e2:
                    err2 = e2.read().decode('utf-8')
                    print(f"    Recipe '{recipe['name']}': ERROR - {err2}")

        # Small delay between batches
        time.sleep(0.2)

    return inserted, errors


def main():
    print(f"Parsing Excel file: {EXCEL_PATH}")
    print(f"Target user: {OWNER_EMAIL}")
    print()

    recipes = parse_excel(EXCEL_PATH)
    print(f"Parsed {len(recipes)} recipes from Excel")
    print()

    # Show categories summary
    categories = {}
    for r in recipes:
        cat = r.get('category', 'Uncategorized') or 'Uncategorized'
        categories[cat] = categories.get(cat, 0) + 1

    print("Categories:")
    for cat, count in sorted(categories.items()):
        print(f"  {cat}: {count} recipes")
    print()

    # Show first few recipes per category for verification
    print("Sample recipes:")
    shown_cats = set()
    for r in recipes:
        cat = r.get('category', '')
        if cat not in shown_cats:
            shown_cats.add(cat)
            ing_count = len(r.get('ingredients', []))
            print(f"  - [{cat}] {r['name']} ({ing_count} ingredients, yield: {r['serving']['total_yield_grams']}g)")
    print(f"  Total: {len(recipes)} recipes")
    print()

    # Ask for confirmation
    if '--dry-run' in sys.argv:
        print("DRY RUN - not inserting into database")
        print()
        for i, r in enumerate(recipes):
            print(f"  {i+1}. [{r['category']}] {r['name']} ({len(r.get('ingredients', []))} ingredients)")
        return

    if '--yes' not in sys.argv:
        response = input(f"Insert {len(recipes)} recipes for {OWNER_EMAIL}? (yes/no): ")
        if response.lower() not in ('yes', 'y'):
            print("Aborted.")
            return

    print(f"Inserting {len(recipes)} recipes...")
    inserted, errors = insert_recipes_batch(recipes)

    print()
    print(f"Done! Inserted: {inserted}, Errors: {errors}")


if __name__ == '__main__':
    main()
