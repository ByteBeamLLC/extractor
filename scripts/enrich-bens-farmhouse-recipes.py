#!/usr/bin/env python3
"""
Enrich imported Ben's Farmhouse recipes with nutritional data from the nutrient Excel.
Only updates recipes that don't already have nutritional information.
"""

import openpyxl
import json
import urllib.request
import urllib.error
import time
import sys

SUPABASE_URL = "https://sijhdlxagliknvhkosjo.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNpamhkbHhhZ2xpa252aGtvc2pvIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MDQ0NDYzMCwiZXhwIjoyMDc2MDIwNjMwfQ.RurV1sZnGm80--rTl5uXG3UoMc061G8fb2ydjmJXCA0"
NUTRIENT_EXCEL = "/Users/talalbazerbachi/Downloads/ben's-farmhouse-2026-02-10-08_22-with-nutrient.xlsx"

# Mapping: Excel column index -> (nutrient display name, unit)
NUTRIENT_COLUMNS = {
    11: ("Energy", "kcal"),
    12: ("Total Carbohydrates", "g"),
    13: ("Protein", "g"),
    14: ("Total Fat", "g"),
    15: ("Dietary Fiber", "g"),
    16: ("Net Carbohydrates", "g"),
    17: ("Calcium", "mg"),
    18: ("Sodium", "mg"),
    19: ("Iron", "mg"),
    20: ("Potassium", "mg"),
    21: ("Saturated Fat", "g"),
    22: ("Trans Fat", "g"),
    23: ("Total Sugar", "g"),
    24: ("Cholesterol", "mg"),
    25: ("Vitamin D", "mcg"),
    26: ("Phosphorus", "mg"),
    27: ("Sugar Alcohol", "g"),
    28: ("Added Sugar", "g"),
    29: ("Salt", "g"),
    30: ("Magnesium", "mg"),
    31: ("Pantothenic acid", "mg"),
    32: ("Polyunsaturated Fat", "g"),
    33: ("Manganese", "mg"),
    34: ("Vitamin A", "mcg"),
    35: ("Vitamin E", "mg"),
    36: ("Vitamin C", "mg"),
    37: ("Vitamin B6", "mg"),
    38: ("Zinc", "mg"),
    39: ("Thiamin", "mg"),
    40: ("Niacin", "mg"),
    41: ("Monounsaturated Fat", "g"),
    42: ("Copper", "mg"),
    43: ("Selenium", "mcg"),
    44: ("Riboflavin", "mg"),
    45: ("Folate", "mcg"),
    46: ("Vitamin B12", "mcg"),
    47: ("Vitamin K", "mcg"),
    48: ("Caffeine", "mg"),
}


def parse_nutrient_excel(filepath):
    """Parse the nutrient Excel and return a dict keyed by (name, category)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Sheet1']
    result = {}

    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row, column=2).value or '').strip()
        category = str(ws.cell(row=row, column=3).value or '').strip()
        sub_category = str(ws.cell(row=row, column=4).value or '').strip()
        status = str(ws.cell(row=row, column=5).value or '').strip()
        serving_weight = ws.cell(row=row, column=6).value
        ingredients_text = ws.cell(row=row, column=7).value
        allergens_text = ws.cell(row=row, column=8).value
        may_contain_text = ws.cell(row=row, column=9).value

        # Parse per-serving nutrients
        per_serving = {}
        for col, (nutrient_name, unit) in NUTRIENT_COLUMNS.items():
            val = ws.cell(row=row, column=col).value
            if val is not None and val != '-' and val != '':
                try:
                    per_serving[nutrient_name] = {
                        "unit": unit,
                        "quantity": round(float(val), 4)
                    }
                except (ValueError, TypeError):
                    pass

        serving_wt = float(serving_weight) if serving_weight else 100

        allergens = []
        if allergens_text and str(allergens_text).strip() not in ('', '-'):
            allergens = [a.strip() for a in str(allergens_text).split(',') if a.strip()]

        may_contain = []
        if may_contain_text and str(may_contain_text).strip() not in ('', '-'):
            may_contain = [a.strip() for a in str(may_contain_text).split(',') if a.strip()]

        ingredient_names = []
        if ingredients_text and str(ingredients_text).strip():
            ingredient_names = [i.strip() for i in str(ingredients_text).split(',') if i.strip()]

        sub_cat = sub_category if sub_category and sub_category != '-' else None

        result[(name, category)] = {
            "serving_weight": serving_wt,
            "per_serving": per_serving,
            "allergens": allergens,
            "may_contain_allergens": may_contain,
            "ingredient_names": ingredient_names,
            "sub_category": sub_cat,
            "status": status,
        }

    return result


def fetch_recipes_needing_nutrition():
    """Fetch all recipes that don't have nutrition per_serving data."""
    url = f'{SUPABASE_URL}/rest/v1/recipes?select=id,name,category,nutrition,serving,status&owner_email=eq.meerim@bensfarmhouse.com&limit=1000'
    req = urllib.request.Request(url)
    req.add_header('apikey', SERVICE_ROLE_KEY)
    req.add_header('Authorization', f'Bearer {SERVICE_ROLE_KEY}')
    resp = urllib.request.urlopen(req)
    all_recipes = json.loads(resp.read())

    # Only return recipes with empty/missing per_serving nutrition
    missing = []
    for r in all_recipes:
        nutrition = r.get('nutrition') or {}
        per_serving = nutrition.get('per_serving') or {}
        if not per_serving:
            missing.append(r)

    return missing


def update_recipe(recipe_id, update_data):
    """Update a single recipe in Supabase via PATCH."""
    url = f'{SUPABASE_URL}/rest/v1/recipes?id=eq.{recipe_id}'
    data = json.dumps(update_data).encode('utf-8')
    req = urllib.request.Request(url, data=data, method='PATCH')
    req.add_header('apikey', SERVICE_ROLE_KEY)
    req.add_header('Authorization', f'Bearer {SERVICE_ROLE_KEY}')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Prefer', 'return=minimal')
    urllib.request.urlopen(req)


def build_update_payload(recipe, nutrient_info):
    """Build the PATCH payload for a recipe using nutrient data."""
    per_serving = nutrient_info['per_serving']
    serving_wt = nutrient_info['serving_weight']

    existing_serving = recipe.get('serving') or {}
    total_yield = existing_serving.get('total_yield_grams', 100)
    scale_factor = total_yield / serving_wt if serving_wt > 0 else 1

    # per_recipe_total = per_serving * scale_factor
    per_recipe_total = {}
    for name, val in per_serving.items():
        per_recipe_total[name] = {
            "unit": val["unit"],
            "quantity": round(val["quantity"] * scale_factor, 4)
        }

    calories = per_serving.get('Energy', {}).get('quantity', 0)
    net_carbs = per_serving.get('Net Carbohydrates', {}).get('quantity', 0)

    nutrition = {
        "summary": {"calories": calories, "net_carbs": net_carbs},
        "per_serving": per_serving,
        "per_recipe_total": per_recipe_total,
        "total_yield_grams": total_yield,
        "serving_size_grams": serving_wt
    }

    serving = {
        **existing_serving,
        "serving_size_grams": serving_wt,
        "scale_factor": round(scale_factor, 2),
    }

    is_sub = nutrient_info.get('sub_category') == 'SUB-RECIPE' or recipe.get('category') == 'SUB-RECIPE'

    update = {
        "nutrition": nutrition,
        "serving": serving,
        "allergens": nutrient_info['allergens'],
        "may_contain_allergens": nutrient_info['may_contain_allergens'],
        "status": nutrient_info.get('status', 'IMPORTED'),
        "costs": {
            "food_cost": 0, "net_profit": 0, "total_cost": 0,
            "labour_cost": 0, "vat_percent": 5, "wastage_cost": 0,
            "food_cost_type": "AED", "packaging_cost": 0,
            "labour_cost_type": "AED", "wastage_cost_type": "percentage",
            "selling_price_exc_vat": 0, "selling_price_inc_vat": 0,
            "total_ingredient_cost": 0,
        },
        "labels": {
            "allergen_list": nutrient_info['allergens'],
            "diet_type_list": [],
            "ingredient_list": nutrient_info['ingredient_names'],
            "business_address_list": []
        },
        "metadata": {
            "qr_image": None, "cover_image": None,
            "is_sub_recipe": is_sub,
            "packaging_logo": None, "is_imported_recipe": True
        },
    }

    if nutrient_info.get('sub_category'):
        update['sub_category'] = nutrient_info['sub_category']

    return update


def main():
    print(f"Reading nutrient data from: {NUTRIENT_EXCEL}")
    nutrient_data = parse_nutrient_excel(NUTRIENT_EXCEL)
    print(f"Parsed {len(nutrient_data)} recipes with nutrient data")

    print(f"\nFetching recipes without nutrition from Supabase...")
    recipes = fetch_recipes_needing_nutrition()
    print(f"Found {len(recipes)} recipes missing nutrition data")

    # Match by (name, category), fallback to name-only
    matched = []
    unmatched = []
    for r in recipes:
        name = r['name']
        category = r['category'] or ''
        key = (name, category)
        if key in nutrient_data:
            matched.append((r, nutrient_data[key]))
        else:
            found = False
            for (n, c), ndata in nutrient_data.items():
                if n == name:
                    matched.append((r, ndata))
                    found = True
                    break
            if not found:
                unmatched.append(r)

    print(f"\nMatched: {len(matched)} recipes")
    print(f"Unmatched: {len(unmatched)} recipes")

    if unmatched:
        print("\nUnmatched recipes (no nutrient data available):")
        for r in unmatched[:20]:
            print(f"  - [{r['category']}] {r['name']}")
        if len(unmatched) > 20:
            print(f"  ... and {len(unmatched) - 20} more")

    if not matched:
        print("No recipes to update.")
        return

    sample_recipe, sample_nutrient = matched[0]
    sample_update = build_update_payload(sample_recipe, sample_nutrient)
    print(f"\nSample update for '{sample_recipe['name']}':")
    print(f"  Calories: {sample_update['nutrition']['summary']['calories']}")
    print(f"  Allergens: {sample_update['allergens']}")
    print(f"  May contain: {sample_update['may_contain_allergens']}")
    print(f"  Nutrients ({len(sample_update['nutrition']['per_serving'])}): {list(sample_update['nutrition']['per_serving'].keys())[:10]}...")
    print(f"  Status: {sample_update['status']}")

    if '--dry-run' in sys.argv:
        print("\nDRY RUN - not updating database")
        return

    if '--yes' not in sys.argv:
        response = input(f"\nUpdate {len(matched)} recipes with nutrition data? (yes/no): ")
        if response.lower() not in ('yes', 'y'):
            print("Aborted.")
            return

    print(f"\nUpdating {len(matched)} recipes...")
    updated = 0
    errors = 0

    for i, (recipe, nutrient_info) in enumerate(matched):
        try:
            update_data = build_update_payload(recipe, nutrient_info)
            update_recipe(recipe['id'], update_data)
            updated += 1
            if (i + 1) % 50 == 0:
                print(f"  Progress: {i + 1}/{len(matched)} updated")
                time.sleep(0.1)
        except urllib.error.HTTPError as e:
            err = e.read().decode('utf-8')
            print(f"  ERROR updating '{recipe['name']}': {e.code} - {err}")
            errors += 1

    print(f"\nDone! Updated: {updated}, Errors: {errors}")


if __name__ == '__main__':
    main()
